"""
UGBS Research NLP Analyser — Streamlit App
Author: Jonathan Elikem (University of Ghana Business School)
"""

import os, re, random, warnings, io
import requests
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import seaborn as sns
import streamlit as st
from collections import Counter
from wordcloud import WordCloud

warnings.filterwarnings("ignore")

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="UGBS Research NLP Analyser",
    page_icon="🎓",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Theme CSS ─────────────────────────────────────────────────────────────────
st.markdown("""
<style>
/* ── Base ── */
[data-testid="stAppViewContainer"] { background: #F4F6FB; }
[data-testid="stSidebar"] { background: #1E2761 !important; }
[data-testid="stSidebar"] * { color: #CADCFC !important; }
[data-testid="stSidebar"] .stSelectbox label,
[data-testid="stSidebar"] .stTextInput label { color: #CADCFC !important; }
[data-testid="stSidebarNav"] { display: none; }

/* ── Top banner ── */
.top-banner {
    background: linear-gradient(135deg, #1E2761 0%, #2E4A8F 60%, #0D9488 100%);
    padding: 1.8rem 2rem 1.4rem;
    border-radius: 14px;
    margin-bottom: 1.5rem;
}
.top-banner h1 { color: white; font-size: 1.9rem; margin: 0; font-weight: 700; }
.top-banner p  { color: #CADCFC; margin: 0.3rem 0 0; font-size: 0.95rem; }
.creator-tag   { color: #E8A020 !important; font-weight: 600; }

/* ── Metric cards ── */
.metric-row { display: flex; gap: 1rem; margin-bottom: 1.2rem; flex-wrap: wrap; }
.metric-card {
    flex: 1; min-width: 130px;
    border-radius: 12px;
    padding: 1rem 1.2rem;
    text-align: center;
    color: white;
}
.metric-card .num { font-size: 2.2rem; font-weight: 700; line-height: 1; }
.metric-card .lbl { font-size: 0.78rem; opacity: 0.88; margin-top: 4px; }
.mc-navy   { background: #1E2761; }
.mc-teal   { background: #0D9488; }
.mc-purple { background: #7C3AED; }
.mc-amber  { background: #D97706; }
.mc-coral  { background: #E05A2B; }
.mc-green  { background: #059669; }

/* ── Section headers ── */
.sec-header {
    font-size: 1.15rem; font-weight: 700;
    padding: 0.5rem 0.9rem;
    border-radius: 8px;
    margin: 1.2rem 0 0.8rem;
    color: white;
}
.sh-navy   { background: #1E2761; }
.sh-teal   { background: #0D9488; }
.sh-purple { background: #7C3AED; }
.sh-amber  { background: #D97706; }
.sh-coral  { background: #E05A2B; }
.sh-green  { background: #059669; }
.sh-pink   { background: #9D174D; }

/* ── Paper card ── */
.paper-card {
    background: white;
    border-radius: 10px;
    border-left: 5px solid #1E2761;
    padding: 0.9rem 1rem;
    margin-bottom: 0.7rem;
    box-shadow: 0 1px 4px rgba(0,0,0,0.06);
}
.paper-card .ptitle  { font-weight: 700; font-size: 0.9rem; color: #1E2761; }
.paper-card .pabstr  { font-size: 0.82rem; color: #555; margin-top: 4px; }
.paper-card .psumm   { font-size: 0.82rem; color: #374151; background: #F0F9FF;
                        border-radius: 6px; padding: 6px 8px; margin-top: 6px; }
.paper-card .pmeta   { font-size: 0.75rem; color: #9CA3AF; margin-top: 5px; }

/* ── Accuracy badge ── */
.badge-good { background:#D1FAE5; color:#065F46; border-radius:6px;
              padding:2px 8px; font-size:0.78rem; font-weight:600; }
.badge-acc  { background:#FEF3C7; color:#92400E; border-radius:6px;
              padding:2px 8px; font-size:0.78rem; font-weight:600; }
.badge-poor { background:#FEE2E2; color:#991B1B; border-radius:6px;
              padding:2px 8px; font-size:0.78rem; font-weight:600; }

/* ── Tab styling ── */
.stTabs [data-baseweb="tab-list"] { gap: 6px; }
.stTabs [data-baseweb="tab"] {
    border-radius: 8px 8px 0 0 !important;
    font-weight: 600 !important;
    font-size: 0.88rem !important;
}

/* ── Footer ── */
.footer {
    text-align: center; padding: 1.5rem;
    color: #9CA3AF; font-size: 0.8rem;
    border-top: 1px solid #E5E7EB;
    margin-top: 2rem;
}
</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
#  NLP IMPORTS (deferred to avoid blocking page load)
# ══════════════════════════════════════════════════════════════════════════════
@st.cache_resource(show_spinner="Loading NLP libraries…")
def load_nlp_libs():
    import nltk
    for pkg in ["punkt", "stopwords", "wordnet", "averaged_perceptron_tagger", "punkt_tab"]:
        try:
            nltk.download(pkg, quiet=True)
        except Exception:
            pass
    from nltk.corpus   import stopwords as sw
    from nltk.tokenize import word_tokenize, sent_tokenize
    from nltk.stem     import WordNetLemmatizer
    from sklearn.feature_extraction.text import CountVectorizer, TfidfVectorizer
    from sklearn.decomposition           import LatentDirichletAllocation
    try:
        from sentence_transformers import SentenceTransformer
        embedding_model = SentenceTransformer("all-MiniLM-L6-v2")
    except ImportError:
        SentenceTransformer = None
        embedding_model = None
    # BERTopic requires hdbscan/umap which don't install cleanly on all platforms
    # The app runs fine with LDA only when BERTopic is unavailable
    try:
        from bertopic import BERTopic
    except (ImportError, Exception):
        BERTopic = None

    STOP_WORDS = set(sw.words("english"))
    STOP_WORDS.update({
        "abstract","introduction","conclusion","discussion","methodology",
        "section","chapter","figure","table","appendix","fig","et","al",
        "study","paper","research","article","examine","explore","investigate",
        "analyze","analyse","assess","evaluate","review","propose","present",
        "demonstrate","show","find","found","suggest","indicate",
        "result","results","finding","findings","effect","impact",
        "significant","positive","negative","relationship","variable","variables",
        "analysis","analytical","approach","method","methodology",
        "model","models","framework","based","using","used","use",
        "also","however","therefore","thus","hence","moreover","furthermore",
        "although","while","whereas","among","within","across","towards",
        "new","high","low","large","small","key","main","major","important",
        "different","various","several","many","much","less","well",
        "may","might","could","would","should","shall","must",
        "one","two","three","four","five","first","second","third",
    })
    lemmatizer = WordNetLemmatizer()
    # embedding_model loaded inside try/except above
    if 'embedding_model' not in dir():
        embedding_model = None
    return (STOP_WORDS, lemmatizer, word_tokenize, sent_tokenize,
            CountVectorizer, TfidfVectorizer, LatentDirichletAllocation,
            embedding_model, BERTopic)


@st.cache_resource(show_spinner="Loading BART model (~1.6 GB on first run)…")
def load_bart():
    try:
        from transformers import AutoTokenizer, AutoModelForSeq2SeqLM
        import torch
        name = "facebook/bart-large-cnn"
        tok  = AutoTokenizer.from_pretrained(name)
        mdl  = AutoModelForSeq2SeqLM.from_pretrained(name)
        mdl.eval()
        return tok, mdl
    except Exception as e:
        st.error(f"Could not load BART model: {e}")
        return None, None


# ══════════════════════════════════════════════════════════════════════════════
#  TEXT UTILITIES
# ══════════════════════════════════════════════════════════════════════════════
def strip_front_matter(text):
    m = re.search(r"\bAbstract\b", text[:2000], re.IGNORECASE)
    if m:
        return text[m.start():]
    return text[min(500, max(300, len(text)//20)):]

def get_full_text(raw):
    t = strip_front_matter(raw)
    t = re.sub(r"\n\s*\d+\s*\n", "\n", t)
    t = re.sub(r"\bReferences?\s*\n.*", "", t, flags=re.DOTALL|re.IGNORECASE)
    t = re.sub(r"\[\d+[,\d\s]*\]", "", t)
    t = re.sub(r"http\S+", "", t)
    t = re.sub(r"[^\w\s.,;:()\-]", " ", t)
    return re.sub(r"\s+", " ", t).strip()

def extract_abstract(raw):
    m = re.search(
        r"(?i)\ba\s*b\s*s\s*t\s*r\s*a\s*c\s*t\b[\s:–—]*(.*?)"
        r"(?=\n\s*\n|\b(?:keywords?|introduction|1\.|©)\b)",
        raw, re.DOTALL)
    if m:
        return re.sub(r"\s+", " ", m.group(1)).strip()[:4000]
    return re.sub(r"\s+", " ", raw[:600]).strip()

def extract_year(raw):
    s = raw[:3000]
    for pat in [r"[©(c)]\s*(19|20)\d{2}",
                r"(?:published|received|accepted)[^\d]{0,20}(19|20)\d{2}",
                r"vol\.?\s*\d+.{0,30}(19|20)\d{2}"]:
        m = re.search(pat, s, re.IGNORECASE)
        if m:
            return int(re.search(r"(19|20)\d{2}", m.group()).group())
    years = re.findall(r"\b(19[9]\d|20[0-3]\d)\b", s)
    return int(years[0]) if years else 0

MAX_FNAME = 80
def safe_fname(name):
    if len(name) <= MAX_FNAME:
        return name
    stem, ext = os.path.splitext(name)
    return stem[:MAX_FNAME - len(ext) - 5] + ext


# ══════════════════════════════════════════════════════════════════════════════
#  CORPUS LOADER
# ══════════════════════════════════════════════════════════════════════════════
def parse_paper(raw, display_name):
    return {
        "title":     os.path.splitext(os.path.basename(display_name))[0],
        "abstract":  extract_abstract(raw),
        "full_text": get_full_text(raw),
        "raw_text":  raw,
        "path":      display_name,
        "year":      extract_year(raw),
    }

def load_local(root):
    if not os.path.isdir(root):
        return None
    hod_dirs = sorted(d for d in os.listdir(root)
                      if os.path.isdir(os.path.join(root, d)) and not d.startswith("."))
    if not hod_dirs:
        return None
    corpus = {}
    for hod in hod_dirs:
        hod_path = os.path.join(root, hod)
        txts = sorted(f for f in os.listdir(hod_path) if f.lower().endswith(".txt"))
        corpus[hod] = []
        for fname in txts:
            try:
                with open(os.path.join(hod_path, fname), encoding="utf-8", errors="replace") as fh:
                    corpus[hod].append(parse_paper(fh.read(), fname))
            except Exception:
                pass
    return corpus

def load_github(root, api_base, raw_base, folder, progress_cb=None):
    r = requests.get(f"{api_base}/{folder}", timeout=30)
    r.raise_for_status()
    hod_folders = [x for x in r.json() if x["type"] == "dir"]
    os.makedirs(root, exist_ok=True)
    corpus = {}
    total  = len(hod_folders)
    for fi, hod_item in enumerate(sorted(hod_folders, key=lambda x: x["name"])):
        hod_name = hod_item["name"]
        hod_dir  = os.path.join(root, hod_name)
        os.makedirs(hod_dir, exist_ok=True)
        r2 = requests.get(f"{api_base}/{hod_item['path']}", timeout=30)
        r2.raise_for_status()
        txts = sorted([f for f in r2.json() if f["name"].lower().endswith(".txt")],
                      key=lambda x: x["name"])
        corpus[hod_name] = []
        used = set()
        for f in txts:
            try:
                raw = requests.get(f"{raw_base}/{f['path']}", timeout=30).text
                sf  = safe_fname(f["name"])
                base, ext = os.path.splitext(sf)
                cand, n = sf, 1
                while cand in used:
                    cand = f"{base[:MAX_FNAME-len(ext)-5]}_{n}{ext}"
                    n   += 1
                used.add(cand)
                with open(os.path.join(hod_dir, cand), "w", encoding="utf-8") as fh:
                    fh.write(raw)
                corpus[hod_name].append(parse_paper(raw, f["name"]))
            except Exception:
                pass
        if progress_cb:
            progress_cb((fi + 1) / total, hod_name)
    return corpus


# ══════════════════════════════════════════════════════════════════════════════
#  NLP FUNCTIONS
# ══════════════════════════════════════════════════════════════════════════════
def get_preprocess_fn(STOP_WORDS, lemmatizer, word_tokenize):
    def preprocess(text, min_len=3):
        return [lemmatizer.lemmatize(t) for t in word_tokenize(text.lower())
                if t.isalpha() and len(t) >= min_len and t not in STOP_WORDS]
    return preprocess

def extractive_summary(text, n=5, TfidfVectorizer=None, sent_tokenize=None):
    sentences = sent_tokenize(text)
    if len(sentences) <= n:
        return " ".join(sentences)
    vec = TfidfVectorizer(stop_words="english", max_features=500)
    try:
        m      = vec.fit_transform(sentences)
        scores = np.asarray(m.sum(axis=1)).flatten()
        idx    = sorted(np.argsort(scores)[-n:].tolist())
        return " ".join([sentences[i] for i in idx])
    except Exception:
        return " ".join(sentences[:n])

def t5_summarise(text, tok, mdl, max_chars=3000):
    if not text or not text.strip() or tok is None or mdl is None:
        return ""
    try:
        import torch
        inp = tok(text[:max_chars].strip(), return_tensors="pt",
                  max_length=1024, truncation=True)
        with torch.no_grad():
            ids = mdl.generate(inp["input_ids"], max_length=120, min_length=30,
                               num_beams=4, early_stopping=True, no_repeat_ngram_size=3)
        return tok.decode(ids[0], skip_special_tokens=True)
    except Exception:
        return ""

def run_lda(papers, STOP_WORDS, CountVectorizer, LatentDirichletAllocation,
            n_topics=5, n_words=8):
    texts  = [p["full_text"] for p in papers]
    actual = min(n_topics, max(2, len(papers)-1))
    vec    = CountVectorizer(max_df=0.95, min_df=1, stop_words=list(STOP_WORDS),
                             max_features=1000)
    dtm    = vec.fit_transform(texts)
    fn     = vec.get_feature_names_out()
    lda    = LatentDirichletAllocation(n_components=actual, max_iter=20,
                                       learning_method="online", random_state=42)
    lda.fit(dtm)
    topic_words = [[fn[i] for i in t.argsort()[:-n_words-1:-1]] for t in lda.components_]
    dominant    = lda.transform(dtm).argmax(axis=1).tolist()
    return topic_words, dominant

def run_bertopic(papers, STOP_WORDS, embedding_model, BERTopic, CountVectorizer):
    if embedding_model is None or BERTopic is None:
        return {}, []
    texts = [p.get("full_text") or p.get("abstract","") for p in papers]
    texts = [t for t in texts if isinstance(t,str) and len(t.split())>50]
    if len(texts) < 2:
        return {}, []
    vec_model = CountVectorizer(stop_words=list(STOP_WORDS), ngram_range=(1,2),
                                min_df=1, max_df=1.0)
    min_size  = max(2, len(texts)//10)
    try:
        model      = BERTopic(embedding_model=embedding_model,
                               vectorizer_model=vec_model,
                               min_topic_size=min_size, verbose=False)
        topics, _  = model.fit_transform(texts)
        info       = model.get_topic_info()
        kws        = {}
        for tid in info["Topic"].tolist():
            if tid != -1:
                kw = model.get_topic(tid)
                if kw:
                    kws[tid] = [w for w, _ in kw[:8]]
        return kws, topics
    except Exception:
        return {}, []

# ── Validation ────────────────────────────────────────────────────────────────
DOMAIN_WORDS = {
    "digital","social","media","government","data","health","supply","chain",
    "risk","market","consumer","leadership","performance","innovation",
    "sustainability","corporate","financial","bank","credit","governance",
    "policy","public","technology","information","system","framework",
    "education","environment","gender","income","poverty","growth","trade",
    "energy","climate","rural","urban","community","hospital","patient",
}

def keyword_overlap(abstract, summary, preprocess):
    a = set(preprocess(abstract)); s = set(preprocess(summary))
    return round(len(a & s) / len(a) * 100, 1) if a else 0.0

def rouge1_f1(abstract, summary, preprocess):
    ref = set(preprocess(abstract)); hyp = set(preprocess(summary))
    if not ref or not hyp:
        return 0.0
    ov = len(ref & hyp)
    p  = ov / len(hyp); r = ov / len(ref)
    return round(2*p*r/(p+r)*100, 1) if (p+r) else 0.0

def detect_issues(abstract, summary, full_text, preprocess, abstractive=False):
    issues = []
    at = set(preprocess(abstract)); st = set(preprocess(summary))
    missed = (at - st) & DOMAIN_WORDS
    if missed:
        issues.append(f"Missing concepts: {', '.join(sorted(missed)[:4])}")
    if len(summary.split()) < (15 if abstractive else 30):
        issues.append("Summary too short")
    halluc = False
    if abstractive:
        ft  = set(preprocess(full_text))
        ph  = st - ft - DOMAIN_WORDS
        halluc = len(ph) > 6
        if halluc:
            issues.append(f"Possible hallucination ({len(ph)} phantom words)")
        neg = {"not","no","never","without","lack","fail","poor","low","decline"}
        ex  = set(summary.lower().split()) & neg - set(abstract.lower().split()) & neg
        if ex:
            issues.append(f"Possible misrepresentation ({', '.join(list(ex)[:3])})")
    if not issues:
        issues.append("None detected")
    return issues, halluc

def rating_label(ov):
    if ov >= 60: return "✅ Good", "badge-good"
    if ov >= 35: return "⚠️ Acceptable", "badge-acc"
    return "❌ Poor", "badge-poor"


# ══════════════════════════════════════════════════════════════════════════════
#  ANALYSIS SIGNALS
# ══════════════════════════════════════════════════════════════════════════════
THEORY_KW  = ["theory","theoretical","conceptual","framework","literature","construct",
               "model","paradigm","hypothesis","proposition","axiom","epistemolog",
               "ontolog","philosophi","discours","notion","concept"]
APPLIED_KW = ["practical","implication","recommend","industry","practitioner","manag",
               "policy","decision","implement","real","world","business","operati",
               "case","problem","solution","strateg","intervention","program",
               "stakeholder","government","public","community","patient","hospital"]
QUANT_KW   = ["survey","questionnaire","regression","econometric","quantitative",
               "structural equation","sem","anova","correlation","statistic",
               "measurement","hypothesis","variance","moderation","mediation",
               "panel data","gmm","logit","probit","covariance","factor analysis","pls"]
QUAL_KW    = ["interview","qualitative","case study","narrative","ethnograph",
               "grounded theory","thematic","discourse","observation","focus group",
               "phenomenolog","content analysis","in-depth","exploratory"]
DESIGN_KW  = ["design science","algorithm","prototype","system design","artifact",
               "decision support","fuzzy","topsis","ahp","vikor","mcdm",
               "simulation","framework development","platform","model development"]
EMERGING   = ["artificial","intelligence","machine","learning","blockchain","digital",
               "transformation","sustainability","esg","fintech","platform","ecosystem",
               "gig","economy","covid","pandemic","remote","hybrid","climate","green",
               "circular","big","analytics","social","media","algorithm","automation",
               "open","government"]


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL EXPORT
# ══════════════════════════════════════════════════════════════════════════════
def build_excel(results, validation_log, preprocess):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils  import get_column_letter
    wb     = Workbook()
    H_FILL = PatternFill("solid", start_color="1E2761")
    A_FILL = PatternFill("solid", start_color="D6E4F0")
    W_FILL = PatternFill("solid", start_color="FFFFFF")
    H_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    B_FONT = Font(name="Arial", size=10)
    WRAP   = Alignment(wrap_text=True, vertical="top")
    CTR    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    THIN   = Border(left=Side(style="thin"),  right=Side(style="thin"),
                    top=Side(style="thin"),   bottom=Side(style="thin"))
    def sc(c, font=None, fill=None, align=None):
        if font:  c.font      = font
        if fill:  c.fill      = fill
        if align: c.alignment = align
        c.border = THIN

    # Sheet 1 — Paper Summaries
    ws1 = wb.active; ws1.title = "Paper Summaries"
    h1  = ["HOD","#","Title","Abstract","Extractive Summary","LDA Keywords","BERTopic","Words"]
    w1  = [28,4,45,60,60,35,35,8]
    ws1.row_dimensions[1].height = 24
    for col,(h,w) in enumerate(zip(h1,w1),1):
        c = ws1.cell(row=1,column=col,value=h); sc(c,H_FONT,H_FILL,CTR)
        ws1.column_dimensions[get_column_letter(col)].width = w
    row = 2
    for hod, hd in results.items():
        for i,(paper,summ) in enumerate(zip(hd["papers"],hd.get("summaries",[])),1):
            fill = A_FILL if row%2==0 else W_FILL
            li   = hd["paper_lda_topic"][i-1] if hd.get("paper_lda_topic") and i-1 < len(hd["paper_lda_topic"]) else 0
            lw   = ", ".join(hd["lda_topics"][li]) if hd.get("lda_topics") and li < len(hd["lda_topics"]) else ""
            bi   = hd["paper_bert_topic"][i-1] if hd.get("paper_bert_topic") and i-1 < len(hd["paper_bert_topic"]) else -1
            bw   = ", ".join(hd.get("bertopic_topics",{}).get(bi,[]))
            vals = [hod,i,paper["title"],paper["abstract"][:600],
                    summ[:600] if summ else "",lw,bw,len(paper["full_text"].split())]
            ws1.row_dimensions[row].height = 70
            for col,v in enumerate(vals,1):
                c = ws1.cell(row=row,column=col,value=v); sc(c,B_FONT,fill,WRAP)
            row += 1
    ws1.freeze_panes = "A2"; ws1.auto_filter.ref = ws1.dimensions

    # Sheet 2 — Validation
    ws2 = wb.create_sheet("Accuracy Validation")
    h2  = ["HOD","Paper","Ext Overlap%","Ext ROUGE-1","Ext Issues","Ext Rating",
           "T5 Overlap%","T5 ROUGE-1","T5 Halluc","T5 Issues","T5 Rating"]
    w2  = [26,42,12,12,40,16,12,12,20,40,16]
    ws2.row_dimensions[1].height = 24
    for col,(h,w) in enumerate(zip(h2,w2),1):
        c = ws2.cell(row=1,column=col,value=h); sc(c,H_FONT,H_FILL,CTR)
        ws2.column_dimensions[get_column_letter(col)].width = w
    row2 = 2
    for hod, val_list in validation_log.items():
        for v in val_list:
            fill = A_FILL if row2%2==0 else W_FILL
            vals = [hod, v["paper"],
                    v.get("ext_overlap",0), v.get("ext_rouge1",0),
                    v.get("ext_issues",""), v.get("ext_rating",""),
                    v.get("t5_overlap",0), v.get("t5_rouge1",0),
                    "⚠️ Possible" if v.get("t5_halluc") else "✅ None",
                    v.get("t5_issues",""), v.get("t5_rating","")]
            ws2.row_dimensions[row2].height = 42
            for col,val in enumerate(vals,1):
                c = ws2.cell(row=row2,column=col,value=val); sc(c,B_FONT,fill,WRAP)
            row2 += 1
    ws2.freeze_panes = "A2"

    # Sheet 3 — T5 Summaries
    ws3 = wb.create_sheet("T5 Summaries")
    for col,(h,w) in enumerate(zip(["HOD","#","Title","T5 Summary"],[28,4,45,90]),1):
        c = ws3.cell(row=1,column=col,value=h); sc(c,H_FONT,H_FILL,CTR)
        ws3.column_dimensions[get_column_letter(col)].width = w
    r3 = 2
    for hod, hd in results.items():
        for i,(paper,t5s) in enumerate(zip(hd["papers"],hd.get("t5_summaries",[])),1):
            fill = A_FILL if r3%2==0 else W_FILL
            for col,v in enumerate([hod,i,paper["title"],t5s],1):
                c = ws3.cell(row=r3,column=col,value=v); sc(c,B_FONT,fill,WRAP)
            ws3.row_dimensions[r3].height = 60; r3 += 1
    ws3.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
#  SIDEBAR
# ══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("## ⚙️ Configuration")
    st.markdown("---")
    st.markdown("**GitHub Settings**")
    gh_user   = st.text_input("GitHub Username", value="jonathanelikem")
    gh_repo   = st.text_input("Repository Name", value="Automatic_Summarization_NLP")
    gh_branch = st.text_input("Branch", value="main")
    gh_folder = st.text_input("Corpus Folder", value="HOD_TXT_CORPUS")
    st.markdown("---")
    st.markdown("**NLP Settings**")
    n_topics     = st.slider("LDA Topics per HOD", 2, 8, 5)
    n_val_papers = st.slider("Validation papers per HOD", 1, 5, 3)
    run_bart     = st.checkbox("Run BART (T5) summaries", value=True,
                               help="Disable to skip the ~1.6 GB model download")
    st.markdown("---")
    st.markdown("**Local Corpus**")
    local_override = st.text_input("Override local path (optional)", value="",
                                   placeholder="C:\\Users\\...\\HOD_TXT_CORPUS")
    st.markdown("---")
    st.markdown("""
    <div style='text-align:center; padding-top:1rem;'>
        <div style='font-size:0.7rem; color:#CADCFC; opacity:0.7;'>
            Created by<br>
            <span style='font-size:0.85rem; color:#E8A020; font-weight:700;'>
                Jonathan Elikem
            </span><br>
            UG Business School · 2026
        </div>
    </div>""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
#  BANNER
# ══════════════════════════════════════════════════════════════════════════════
st.markdown("""
<div class="top-banner">
  <h1>🎓 UGBS Research NLP Analyser</h1>
  <p>Automatic text summarization &amp; cross-HOD research theme analysis ·
     <span class="creator-tag">Jonathan Elikem</span> ·
     University of Ghana Business School</p>
</div>""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
#  SESSION STATE
# ══════════════════════════════════════════════════════════════════════════════
for key in ["corpus","results","validation_log","nlp_libs","bart_loaded"]:
    if key not in st.session_state:
        st.session_state[key] = None

GH_API  = f"https://api.github.com/repos/{gh_user}/{gh_repo}/contents"
GH_RAW  = f"https://raw.githubusercontent.com/{gh_user}/{gh_repo}/{gh_branch}"
LOCAL_ROOT = (local_override.strip() if local_override.strip()
              else os.path.join(os.getcwd(), gh_folder))


# ══════════════════════════════════════════════════════════════════════════════
#  TABS
# ══════════════════════════════════════════════════════════════════════════════
tabs = st.tabs(["🏠 Home", "📚 Corpus", "🤖 NLP Processing",
                "✅ Validation", "📊 Analysis", "💾 Export"])


# ─────────────────────────────────────────────────────────────────────────────
# TAB 0 — HOME
# ─────────────────────────────────────────────────────────────────────────────
with tabs[0]:
    col1, col2, col3, col4 = st.columns(4)
    corpus_size = sum(len(v) for v in st.session_state.corpus.values()) if st.session_state.corpus else 0
    n_hods      = len(st.session_state.corpus) if st.session_state.corpus else 0
    n_results   = len(st.session_state.results) if st.session_state.results else 0
    n_val       = sum(len(v) for v in st.session_state.validation_log.values()) if st.session_state.validation_log else 0

    for col, (num, lbl, cls) in zip(
        [col1, col2, col3, col4],
        [(n_hods,"HODs Loaded","mc-navy"),
         (corpus_size,"Papers Loaded","mc-teal"),
         (n_results,"HODs Processed","mc-purple"),
         (n_val,"Papers Validated","mc-amber")]
    ):
        with col:
            st.markdown(f"""<div class="metric-card {cls}">
                <div class="num">{num}</div>
                <div class="lbl">{lbl}</div></div>""", unsafe_allow_html=True)

    st.markdown('<div class="sec-header sh-navy">📋 Project Overview</div>', unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    with c1:
        st.markdown("""
**This app automates three key tasks:**

1. **Corpus loading** — pulls `.txt` paper files from a GitHub repo (or local disk) organised by HOD folder
2. **NLP Summarisation** — runs LDA, BERTopic, and BART across all papers per HOD
3. **Research Analysis** — answers 6 cross-departmental questions about UGBS research identity
        """)
    with c2:
        st.markdown("""
**How to run (in order):**

1. ⚙️ Configure GitHub details in the sidebar
2. 📚 Go to **Corpus** tab → Load corpus
3. 🤖 Go to **NLP Processing** → Run pipeline
4. ✅ Go to **Validation** → Run accuracy checks
5. 📊 Go to **Analysis** → Explore all 6 questions
6. 💾 Go to **Export** → Download Excel
        """)

    st.markdown('<div class="sec-header sh-teal">🤖 NLP Models</div>', unsafe_allow_html=True)
    mc1, mc2, mc3 = st.columns(3)
    for col, (name, typ, desc, cls) in zip(
        [mc1, mc2, mc3],
        [("LDA", "Extractive · Topic Modelling",
          "Groups words into latent topics. Fast, interpretable keyword clusters.", "sh-navy"),
         ("BERTopic", "Extractive · Semantic Clustering",
          "Encodes papers with BERT embeddings and clusters by semantic similarity.", "sh-purple"),
         ("BART (T5)", "Abstractive · Text Generation",
          "Reads paper text and generates brand-new summary sentences.", "sh-teal")]
    ):
        with col:
            st.markdown(f'<div class="sec-header {cls}">{name} — {typ}</div>', unsafe_allow_html=True)
            st.caption(desc)

    st.markdown('<div class="sec-header sh-amber">📖 Pipeline Flow</div>', unsafe_allow_html=True)
    st.code("""
Load corpus (local / GitHub)
       ↓
Clean text (strip noise, front matter, references)
       ↓
Extract summaries: TF-IDF extractive · LDA · BERTopic · BART
       ↓
Accuracy validation: Keyword Overlap % · ROUGE-1 F1 · Hallucination check
       ↓
Cross-HOD analysis: dominant themes · school of thought ·
                    orientation · methods · hot topics · overlaps
       ↓
Excel export (3 sheets) + visualisations
    """, language="")


# ─────────────────────────────────────────────────────────────────────────────
# TAB 1 — CORPUS
# ─────────────────────────────────────────────────────────────────────────────
with tabs[1]:
    st.markdown('<div class="sec-header sh-teal">📚 Load Corpus</div>', unsafe_allow_html=True)

    src_status = ""
    if os.path.isdir(LOCAL_ROOT):
        st.success(f"✅ Local corpus found at: `{LOCAL_ROOT}` — loading from disk will be instant.")
        src_status = "local"
    else:
        st.info(f"📡 No local corpus at `{LOCAL_ROOT}`. Will download from GitHub and save locally.")
        src_status = "github"

    if st.button("🚀 Load Corpus", type="primary", use_container_width=True):
        with st.spinner("Loading corpus…"):
            try:
                if os.path.isdir(LOCAL_ROOT):
                    corpus = load_local(LOCAL_ROOT)
                    if corpus:
                        st.session_state.corpus = corpus
                        st.success(f"Loaded {sum(len(v) for v in corpus.values())} papers from {len(corpus)} HODs  [LOCAL]")
                    else:
                        st.warning("Folder exists but is empty — trying GitHub…")
                        src_status = "github"

                if src_status == "github" or not st.session_state.corpus:
                    prog = st.progress(0, text="Connecting to GitHub…")
                    def cb(frac, name):
                        prog.progress(frac, text=f"Downloading: {name}")
                    corpus = load_github(LOCAL_ROOT, GH_API, GH_RAW, gh_folder, cb)
                    st.session_state.corpus = corpus
                    prog.empty()
                    st.success(f"Downloaded & saved {sum(len(v) for v in corpus.values())} papers from {len(corpus)} HODs  [GITHUB → LOCAL]")
            except Exception as e:
                st.error(f"❌ Failed to load corpus: {e}")

    if st.session_state.corpus:
        st.markdown('<div class="sec-header sh-teal">📂 Corpus Contents</div>', unsafe_allow_html=True)
        for hod, papers in st.session_state.corpus.items():
            with st.expander(f"📂 {hod}  ({len(papers)} papers)", expanded=False):
                for p in papers:
                    yr = f" · {p['year']}" if p.get("year") else ""
                    st.markdown(f"""<div class="paper-card">
                        <div class="ptitle">{p['title'][:90]}</div>
                        <div class="pabstr">{p['abstract'][:250]}…</div>
                        <div class="pmeta">{len(p['full_text'].split())} words{yr}</div>
                    </div>""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# TAB 2 — NLP PROCESSING
# ─────────────────────────────────────────────────────────────────────────────
with tabs[2]:
    st.markdown('<div class="sec-header sh-purple">🤖 NLP Processing</div>', unsafe_allow_html=True)

    if not st.session_state.corpus:
        st.warning("Load the corpus first (📚 Corpus tab).")
    else:
        if st.button("▶️ Run Full NLP Pipeline", type="primary", use_container_width=True):
            # Load libraries
            with st.spinner("Loading NLP libraries…"):
                (STOP_WORDS, lemmatizer, word_tokenize, sent_tokenize,
                 CountVectorizer, TfidfVectorizer, LDA,
                 embedding_model, BERTopic_cls) = load_nlp_libs()
                preprocess = get_preprocess_fn(STOP_WORDS, lemmatizer, word_tokenize)
                st.session_state["preprocess"] = preprocess
                st.session_state["STOP_WORDS"] = STOP_WORDS

            if run_bart:
                with st.spinner("Loading BART model…"):
                    bart_tok, bart_mdl = load_bart()

            results      = {}
            total_hods   = len(st.session_state.corpus)
            overall_prog = st.progress(0, text="Starting pipeline…")
            log_area     = st.empty()

            random.seed(42)
            for hi, (hod_name, papers) in enumerate(st.session_state.corpus.items()):
                log_area.info(f"Processing: **{hod_name}** ({len(papers)} papers)…")
                hd = {
                    "papers":           papers,
                    "summaries":        [],
                    "t5_summaries":     [],
                    "lda_topics":       [],
                    "paper_lda_topic":  [],
                    "bertopic_topics":  {},
                    "paper_bert_topic": [],
                }
                # Extractive
                for p in papers:
                    hd["summaries"].append(
                        extractive_summary(p["full_text"], 5,
                                           TfidfVectorizer, sent_tokenize))
                # BART
                if run_bart:
                    for p in papers:
                        hd["t5_summaries"].append(
                            t5_summarise(p["full_text"], bart_tok, bart_mdl))
                else:
                    hd["t5_summaries"] = ["" for _ in papers]
                # LDA
                tw, dom = run_lda(papers, STOP_WORDS, CountVectorizer, LDA, n_topics)
                hd["lda_topics"]     = tw
                hd["paper_lda_topic"]= dom
                # BERTopic
                bkws, bdom = run_bertopic(papers, STOP_WORDS, embedding_model, BERTopic_cls, CountVectorizer)
                hd["bertopic_topics"]  = bkws
                hd["paper_bert_topic"] = bdom

                results[hod_name] = hd
                overall_prog.progress((hi+1)/total_hods, text=f"Done: {hod_name}")

            st.session_state.results = results
            log_area.empty()
            overall_prog.empty()
            st.success(f"✅ Pipeline complete — {len(results)} HODs, "
                       f"{sum(len(v['papers']) for v in results.values())} papers processed.")

    if st.session_state.results:
        st.markdown('<div class="sec-header sh-purple">📄 Browse Summaries</div>', unsafe_allow_html=True)
        sel_hod = st.selectbox("Select HOD", list(st.session_state.results.keys()))
        hd      = st.session_state.results[sel_hod]
        for i, (paper, ext_s, t5_s) in enumerate(
            zip(hd["papers"], hd.get("summaries",[]), hd.get("t5_summaries",[])), 1
        ):
            li    = hd["paper_lda_topic"][i-1] if hd.get("paper_lda_topic") and i-1 < len(hd["paper_lda_topic"]) else 0
            lkws  = ", ".join(hd["lda_topics"][li]) if hd.get("lda_topics") and li < len(hd["lda_topics"]) else "—"
            yr    = f" · {paper['year']}" if paper.get("year") else ""
            with st.expander(f"Paper {i}: {paper['title'][:70]}{yr}"):
                c1, c2 = st.columns(2)
                with c1:
                    st.markdown("**Original Abstract**")
                    st.caption(paper["abstract"][:600] + "…")
                    st.markdown("**LDA Topic Keywords**")
                    st.caption(lkws)
                with c2:
                    st.markdown("**Extractive (TF-IDF) Summary**")
                    st.caption(ext_s[:500] + "…" if ext_s else "—")
                    st.markdown("**Abstractive (BART) Summary**")
                    st.caption(t5_s[:400] + "…" if t5_s else "Not generated (disabled in sidebar)")


# ─────────────────────────────────────────────────────────────────────────────
# TAB 3 — VALIDATION
# ─────────────────────────────────────────────────────────────────────────────
with tabs[3]:
    st.markdown('<div class="sec-header sh-coral">✅ Accuracy Validation</div>', unsafe_allow_html=True)

    if not st.session_state.results:
        st.warning("Run NLP pipeline first (🤖 NLP Processing tab).")
    else:
        if st.button("🔍 Run Validation", type="primary", use_container_width=True):
            preprocess = st.session_state.get("preprocess")
            if not preprocess:
                (STOP_WORDS, lemmatizer, word_tokenize, *_) = load_nlp_libs()
                preprocess = get_preprocess_fn(STOP_WORDS, lemmatizer, word_tokenize)
                st.session_state["preprocess"] = preprocess

            random.seed(42)
            val_log = {}
            for hod_name, hd in st.session_state.results.items():
                papers   = hd["papers"]
                ext_sums = hd.get("summaries", [])
                t5_sums  = hd.get("t5_summaries", [])
                sample   = random.sample(range(len(papers)), min(n_val_papers, len(papers)))
                hod_val  = []
                for idx in sample:
                    paper    = papers[idx]
                    ext_s    = ext_sums[idx] if idx < len(ext_sums) else ""
                    t5_s     = t5_sums[idx]  if idx < len(t5_sums)  else ""
                    abst     = paper["abstract"]; ft = paper["full_text"]
                    ext_ov   = keyword_overlap(abst, ext_s, preprocess)
                    ext_r1   = rouge1_f1(abst, ext_s, preprocess)
                    ext_iss, _ = detect_issues(abst, ext_s, ft, preprocess, False)
                    t5_ov    = keyword_overlap(abst, t5_s, preprocess)
                    t5_r1    = rouge1_f1(abst, t5_s, preprocess)
                    t5_iss, t5_h = detect_issues(abst, t5_s, ft, preprocess, True)
                    hod_val.append({
                        "paper":       paper["title"],
                        "ext_overlap": ext_ov, "ext_rouge1": ext_r1,
                        "ext_issues":  " | ".join(ext_iss),
                        "ext_rating":  rating_label(ext_ov)[0],
                        "t5_overlap":  t5_ov, "t5_rouge1": t5_r1,
                        "t5_issues":   " | ".join(t5_iss),
                        "t5_rating":   rating_label(t5_ov)[0],
                        "t5_halluc":   t5_h,
                    })
                val_log[hod_name] = hod_val
            st.session_state.validation_log = val_log
            st.success("✅ Validation complete.")

    if st.session_state.validation_log:
        val_log = st.session_state.validation_log
        # Summary metrics
        st.markdown('<div class="sec-header sh-coral">📊 Validation Summary</div>', unsafe_allow_html=True)

        rows = []
        for hod, vl in val_log.items():
            n  = len(vl)
            rows.append({
                "HOD": hod.split("-")[0].strip()[:28],
                "Papers": n,
                "Ext Overlap %": f"{sum(v['ext_overlap'] for v in vl)/n:.1f}%",
                "Ext ROUGE-1":   f"{sum(v['ext_rouge1']  for v in vl)/n:.1f}%",
                "T5 Overlap %":  f"{sum(v['t5_overlap']  for v in vl)/n:.1f}%",
                "T5 ROUGE-1":    f"{sum(v['t5_rouge1']   for v in vl)/n:.1f}%",
                "Hallucinations":sum(1 for v in vl if v["t5_halluc"]),
            })
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

        # Charts
        st.markdown('<div class="sec-header sh-coral">📈 Overlap Charts</div>', unsafe_allow_html=True)
        flat_hod  = [h.split("-")[0].strip()[:14]+"\n"+v["paper"][:22]
                     for h,vl in val_log.items() for v in vl]
        ext_ovs   = [v["ext_overlap"] for vl in val_log.values() for v in vl]
        t5_ovs    = [v["t5_overlap"]  for vl in val_log.values() for v in vl]

        def bar_chart(vals, labels, title, color_fn):
            fig, ax = plt.subplots(figsize=(max(10, len(labels)*1.5), 7))
            colors  = [color_fn(v) for v in vals]
            bars    = ax.bar(range(len(labels)), vals, color=colors, edgecolor="white")
            for bar, v in zip(bars, vals):
                ax.text(bar.get_x()+bar.get_width()/2, bar.get_height()+0.8,
                        f"{v:.1f}%", ha="center", fontsize=11, fontweight="bold")
            ax.axhline(60, color="#27ae60", linestyle="--", lw=1.5, label="60% threshold")
            ax.axhline(35, color="#e67e22", linestyle="--", lw=1.5, label="35% threshold")
            ax.set_xticks(range(len(labels)))
            ax.set_xticklabels(labels, rotation=30, ha="right", fontsize=9)
            ax.set_ylabel("Overlap (%)", fontsize=14)
            ax.set_title(title, fontsize=16, fontweight="bold", pad=12)
            ax.set_ylim(0, 115)
            ax.legend(fontsize=11)
            ax.spines[["top","right"]].set_visible(False)
            plt.tight_layout()
            return fig

        def ov_color(v):
            return "#2ecc71" if v >= 60 else "#f39c12" if v >= 35 else "#e74c3c"

        c1, c2 = st.columns(2)
        with c1:
            st.pyplot(bar_chart(ext_ovs, flat_hod,
                                "Extractive — Keyword Overlap", ov_color))
        with c2:
            st.pyplot(bar_chart(t5_ovs, flat_hod,
                                "Abstractive (T5/BART) — Keyword Overlap", ov_color))

        # ROUGE-1 grouped bar
        st.markdown('<div class="sec-header sh-coral">📉 ROUGE-1 F1 Comparison</div>', unsafe_allow_html=True)
        hod_ns   = list(val_log.keys())
        ext_r1s  = [sum(v["ext_rouge1"] for v in val_log[h])/len(val_log[h]) for h in hod_ns]
        t5_r1s   = [sum(v["t5_rouge1"]  for v in val_log[h])/len(val_log[h]) for h in hod_ns]
        x = np.arange(len(hod_ns)); w = 0.35
        fig, ax = plt.subplots(figsize=(max(10, len(hod_ns)*2), 7))
        b1 = ax.bar(x-w/2, ext_r1s, w, label="Extractive", color="#2E75B6", edgecolor="white")
        b2 = ax.bar(x+w/2, t5_r1s,  w, label="Abstractive (T5)", color="#E05A2B", edgecolor="white")
        for bar in list(b1)+list(b2):
            if bar.get_height() > 0:
                ax.text(bar.get_x()+bar.get_width()/2, bar.get_height()+0.4,
                        f"{bar.get_height():.1f}", ha="center", fontsize=11, fontweight="bold")
        ax.set_xticks(x)
        ax.set_xticklabels([h.split("-")[0].strip()[:18] for h in hod_ns],
                           rotation=25, ha="right", fontsize=12)
        ax.set_ylabel("ROUGE-1 F1 (%)", fontsize=14)
        ax.set_title("ROUGE-1 F1 — Extractive vs Abstractive per HOD",
                     fontsize=16, fontweight="bold")
        ax.legend(fontsize=12)
        ax.spines[["top","right"]].set_visible(False)
        plt.tight_layout()
        st.pyplot(fig)

        # Detailed per-HOD view
        st.markdown('<div class="sec-header sh-coral">🔎 Detailed Validation Results</div>', unsafe_allow_html=True)
        for hod, vl in val_log.items():
            with st.expander(f"📂 {hod}"):
                for v in vl:
                    rl, cls = rating_label(v["ext_overlap"])
                    rl2, cls2 = rating_label(v["t5_overlap"])
                    st.markdown(f"""
**{v['paper'][:80]}**

| | Extractive | Abstractive (T5) |
|---|---|---|
| Keyword Overlap | {v['ext_overlap']}% | {v['t5_overlap']}% |
| ROUGE-1 F1 | {v['ext_rouge1']}% | {v['t5_rouge1']}% |
| Rating | {rl} | {rl2} |
| Issues | {v['ext_issues']} | {v['t5_issues']} |
| Hallucination | None (extractive) | {"⚠️ Possible" if v['t5_halluc'] else "✅ None"} |
""")
                    st.divider()


# ─────────────────────────────────────────────────────────────────────────────
# TAB 4 — ANALYSIS (questions a–f)
# ─────────────────────────────────────────────────────────────────────────────
with tabs[4]:
    st.markdown('<div class="sec-header sh-green">📊 Cross-HOD Research Analysis</div>', unsafe_allow_html=True)

    if not st.session_state.results:
        st.warning("Run NLP pipeline first (🤖 NLP Processing tab).")
    else:
        results    = st.session_state.results
        preprocess = st.session_state.get("preprocess")
        if not preprocess:
            (STOP_WORDS, lemmatizer, word_tokenize, *_) = load_nlp_libs()
            preprocess = get_preprocess_fn(STOP_WORDS, lemmatizer, word_tokenize)
            st.session_state["preprocess"] = preprocess

        # ── Build keyword index ────────────────────────────────────────────
        all_keywords     = Counter()
        hod_keyword_sets = {}
        for hod, hd in results.items():
            tokens = preprocess(" ".join(p["full_text"] for p in hd["papers"]))
            freq   = Counter(tokens)
            hod_keyword_sets[hod] = set([w for w,_ in freq.most_common(30)])
            all_keywords.update(freq)

        q_tabs = st.tabs(["(a) Themes", "(b) School of Thought",
                          "(c) Orientation", "(d) Methods",
                          "(e) Hot Topics", "(f) Overlaps"])

        # ── (a) Dominant Themes ────────────────────────────────────────────
        with q_tabs[0]:
            st.markdown('<div class="sec-header sh-navy">(a) Dominant Research Themes</div>', unsafe_allow_html=True)
            st.caption("Top keywords and LDA topics across all HODs")

            top30 = all_keywords.most_common(30)
            if top30:
                words, counts = zip(*top30)
                fig, ax = plt.subplots(figsize=(14, 9))
                colors  = plt.cm.get_cmap("tab20", 30)([i/29 for i in range(30)])
                ax.barh(list(reversed(words)), list(reversed(counts)),
                        color=list(reversed(colors)), edgecolor="white", height=0.75)
                ax.set_xlabel("Frequency", fontsize=14)
                ax.set_title("Top 30 Research Keywords — All HODs", fontsize=17, fontweight="bold")
                ax.tick_params(labelsize=13)
                ax.spines[["top","right"]].set_visible(False)
                plt.tight_layout(); st.pyplot(fig)

            # Word clouds
            st.markdown('<div class="sec-header sh-navy">Word Clouds by HOD</div>', unsafe_allow_html=True)
            n_hods  = len(results)
            cols_wc = min(3, n_hods)
            rows_wc = (n_hods + cols_wc - 1) // cols_wc
            CMAPS   = ["Blues","Oranges","Greens","Purples","Reds","YlOrBr","PuBu","BuGn"]
            cols    = st.columns(cols_wc)
            for idx, (hod, hd) in enumerate(results.items()):
                tokens = preprocess(" ".join(p["full_text"] for p in hd["papers"]))
                freq   = Counter(tokens)
                if freq:
                    wc  = WordCloud(width=500, height=300, background_color="white",
                                    colormap=CMAPS[idx % len(CMAPS)],
                                    max_words=50).generate_from_frequencies(freq)
                    fig, ax = plt.subplots(figsize=(5,3))
                    ax.imshow(wc, interpolation="bilinear"); ax.axis("off")
                    ax.set_title(hod.split("-")[0].strip()[:22], fontsize=11, fontweight="bold")
                    plt.tight_layout()
                    with cols[idx % cols_wc]:
                        st.pyplot(fig)

        # ── (b) School of Thought ──────────────────────────────────────────
        with q_tabs[1]:
            st.markdown('<div class="sec-header sh-teal">(b) Shared Intellectual Tradition</div>', unsafe_allow_html=True)
            n_h = len(hod_keyword_sets)
            if n_h > 1:
                shared = set.intersection(*hod_keyword_sets.values())
                union  = set.union(*hod_keyword_sets.values())
                kw_cnt = {w: sum(1 for kws in hod_keyword_sets.values() if w in kws) for w in union}
                near   = {w: c for w,c in kw_cnt.items() if c >= max(2, n_h-1)}

                if shared:
                    st.success(f"**SHARED TRADITION** — {len(shared)} keywords appear in ALL HODs: `{'`, `'.join(sorted(shared))}`")
                elif near:
                    st.info(f"**PARTIAL TRADITION** — {len(near)} keywords shared by most HODs: `{'`, `'.join(sorted(near)[:15])}`")
                else:
                    st.warning("**HIGHLY DIVERSE** — No strong unifying school of thought detected.")

                # Bubble chart
                top25    = sorted(kw_cnt.items(), key=lambda x: (-x[1], -all_keywords[x[0]]))[:25]
                kw_lbs   = [k for k,_ in top25]
                kw_n_hod = [v for _,v in top25]
                kw_freq  = [all_keywords[k] for k in kw_lbs]
                fig, ax  = plt.subplots(figsize=(13,6))
                sc = ax.scatter(range(len(kw_lbs)), kw_n_hod,
                                s=[f/3 for f in kw_freq], c=kw_freq,
                                cmap="YlOrRd", alpha=0.85, edgecolors="grey", lw=0.5)
                plt.colorbar(sc, ax=ax, label="Global frequency", shrink=0.7)
                ax.set_xticks(range(len(kw_lbs)))
                ax.set_xticklabels(kw_lbs, rotation=40, ha="right", fontsize=11)
                ax.set_yticks(range(0, max(kw_n_hod)+2))
                ax.set_ylabel("HODs sharing this keyword", fontsize=14)
                ax.set_title("(b) Shared Intellectual Tradition\n(bubble = frequency; y = HOD count)",
                             fontsize=16, fontweight="bold")
                ax.spines[["top","right"]].set_visible(False)
                plt.tight_layout(); st.pyplot(fig)

        # ── (c) Theoretical vs Applied ────────────────────────────────────
        with q_tabs[2]:
            st.markdown('<div class="sec-header sh-amber">(c) Research Orientation — Alignment with UG Mission</div>', unsafe_allow_html=True)
            orient = {}
            for hod, hd in results.items():
                txt = " ".join(p["full_text"] for p in hd["papers"]).lower()
                w   = txt.split()
                tc  = sum(1 for wd in w if any(k in wd for k in THEORY_KW))
                ac  = sum(1 for wd in w if any(k in wd for k in APPLIED_KW))
                tot = max(tc+ac, 1)
                orient[hod] = {"theory": tc, "applied": ac,
                               "theory_pct":  round(tc/tot*100,1),
                               "applied_pct": round(ac/tot*100,1)}

            short_hods  = [h.split("-")[0].strip()[:18] for h in orient]
            t_pcts      = [orient[h]["theory_pct"]  for h in orient]
            a_pcts      = [orient[h]["applied_pct"] for h in orient]
            x = np.arange(len(short_hods)); w = 0.35

            fig, axes = plt.subplots(1, 2, figsize=(18, 8))
            b1 = axes[0].bar(x-w/2, t_pcts, w, label="Theoretical", color="#8e44ad", edgecolor="white")
            b2 = axes[0].bar(x+w/2, a_pcts, w, label="Applied",     color="#e67e22", edgecolor="white")
            for bar in list(b1)+list(b2):
                if bar.get_height() > 0:
                    axes[0].text(bar.get_x()+bar.get_width()/2, bar.get_height()+0.4,
                                 f"{bar.get_height():.1f}%", ha="center", fontsize=11, fontweight="bold")
            axes[0].set_xticks(x)
            axes[0].set_xticklabels(short_hods, rotation=25, ha="right", fontsize=12)
            axes[0].set_ylabel("Signal share (%)", fontsize=14)
            axes[0].set_title("Theory vs Applied Signal Counts", fontsize=15, fontweight="bold")
            axes[0].legend(fontsize=12); axes[0].spines[["top","right"]].set_visible(False)

            ratios = [orient[h]["applied_pct"]/100 for h in orient]
            clrs   = ["#e67e22" if r>0.55 else "#8e44ad" if r<0.45 else "#2ecc71" for r in ratios]
            axes[1].barh(short_hods, ratios, color=clrs, edgecolor="white", height=0.55)
            axes[1].axvline(0.5, color="grey", ls="--", lw=1.5)
            axes[1].set_xlim(0,1)
            axes[1].set_xlabel("← Theoretical  |  Applied →", fontsize=13)
            axes[1].tick_params(labelsize=12)
            axes[1].set_title("Applied vs Theoretical Balance", fontsize=15, fontweight="bold")
            patches = [mpatches.Patch(color="#e67e22",label="Applied dominant (>55%)"),
                       mpatches.Patch(color="#8e44ad",label="Theory dominant (<45%)"),
                       mpatches.Patch(color="#2ecc71",label="Balanced")]
            axes[1].legend(handles=patches, fontsize=11)
            axes[1].spines[["top","right"]].set_visible(False)
            plt.tight_layout(); st.pyplot(fig)

            avg_applied = sum(ratios)/len(ratios)*100
            if avg_applied > 55:
                st.success(f"**School-wide applied signal share: {avg_applied:.1f}%** — UGBS research is predominantly APPLIED, aligned with the school's professional education mission.")
            else:
                st.info(f"School-wide applied share: {avg_applied:.1f}% — Research is balanced between theory and application.")

        # ── (d) Research Methods ──────────────────────────────────────────
        with q_tabs[3]:
            st.markdown('<div class="sec-header sh-pink">(d) Preferred Research Methods</div>', unsafe_allow_html=True)
            methods = {}
            for hod, hd in results.items():
                txt = " ".join(p["full_text"] for p in hd["papers"]).lower()
                qc  = sum(txt.count(k) for k in QUANT_KW)
                qlc = sum(txt.count(k) for k in QUAL_KW)
                dsc = sum(txt.count(k) for k in DESIGN_KW)
                tot = max(qc+qlc+dsc, 1)
                methods[hod] = {"Quantitative": qc, "Qualitative": qlc, "Design Science": dsc,
                                "q_pct": round(qc/tot*100,1),
                                "ql_pct": round(qlc/tot*100,1),
                                "ds_pct": round(dsc/tot*100,1)}

            hod_list_d = list(methods)
            q_p  = [methods[h]["q_pct"]  for h in hod_list_d]
            ql_p = [methods[h]["ql_pct"] for h in hod_list_d]
            ds_p = [methods[h]["ds_pct"] for h in hod_list_d]
            x    = np.arange(len(hod_list_d)); w = 0.26

            fig, axes = plt.subplots(1, 2, figsize=(20, 9))
            b1 = axes[0].bar(x-w, q_p,  w, label="Quantitative",   color="#2980b9", edgecolor="white")
            b2 = axes[0].bar(x,   ql_p, w, label="Qualitative",    color="#e74c3c", edgecolor="white")
            b3 = axes[0].bar(x+w, ds_p, w, label="Design Science", color="#27ae60", edgecolor="white")
            for bar in list(b1)+list(b2)+list(b3):
                if bar.get_height() > 0.5:
                    axes[0].text(bar.get_x()+bar.get_width()/2, bar.get_height()+0.5,
                                 f"{bar.get_height():.0f}%", ha="center", fontsize=10, fontweight="bold")
            axes[0].set_xticks(x)
            axes[0].set_xticklabels([h.split("-")[0].strip()[:16] for h in hod_list_d],
                                    rotation=25, ha="right", fontsize=12)
            axes[0].set_ylabel("Method signal share (%)", fontsize=14)
            axes[0].set_title("Method Distribution per HOD", fontsize=15, fontweight="bold")
            axes[0].legend(fontsize=12); axes[0].spines[["top","right"]].set_visible(False)

            school_q  = sum(methods[h]["Quantitative"]   for h in methods)
            school_ql = sum(methods[h]["Qualitative"]    for h in methods)
            school_ds = sum(methods[h]["Design Science"] for h in methods)
            tot2 = school_q+school_ql+school_ds
            sizes  = [school_q, school_ql, school_ds]
            labels = ["Quantitative", "Qualitative", "Design Science"]
            colors = ["#2980b9", "#e74c3c", "#27ae60"]
            non_z  = [(s,l,c) for s,l,c in zip(sizes,labels,colors) if s>0]
            vv,ll,cc = zip(*non_z)
            wds,texts,auts = axes[1].pie(vv, labels=ll, colors=cc, autopct="%1.1f%%",
                                          startangle=90, textprops={"fontsize":13},
                                          wedgeprops={"edgecolor":"white","linewidth":2},
                                          pctdistance=0.75)
            for a in auts: a.set_fontsize(13); a.set_fontweight("bold")
            axes[1].set_title("School-Wide Method Distribution", fontsize=15, fontweight="bold")
            plt.tight_layout(); st.pyplot(fig)
            st.info(f"**School-wide:** Quantitative {school_q/tot2*100:.1f}%  ·  Qualitative {school_ql/tot2*100:.1f}%  ·  Design Science {school_ds/tot2*100:.1f}%")

        # ── (e) Emerging Topics ────────────────────────────────────────────
        with q_tabs[4]:
            st.markdown('<div class="sec-header sh-teal">(e) Emerging Hot Topics</div>', unsafe_allow_html=True)
            emerging_data = {}
            for hod, hd in results.items():
                n_recent = max(1, len(hd["papers"])//3)
                recent   = sorted(hd["papers"], key=lambda p: p.get("year",0))[-n_recent:]
                tokens   = set(preprocess(" ".join(p["full_text"] for p in recent)))
                emerging_data[hod] = sorted(tokens & set(EMERGING))
                if emerging_data[hod]:
                    st.markdown(f"**{hod.split('-')[0].strip()}:** `{'` · `'.join(emerging_data[hod])}`")
                else:
                    st.markdown(f"**{hod.split('-')[0].strip()}:** No clear emerging signals")

            all_em = sorted({s for sigs in emerging_data.values() for s in sigs})
            if all_em and len(results) >= 2:
                hod_ns   = list(emerging_data.keys())
                mat      = np.array([[1 if sig in emerging_data[h] else 0 for sig in all_em] for h in hod_ns])
                fig, ax  = plt.subplots(figsize=(max(10,len(all_em)*0.9), max(4,len(hod_ns)*0.9)))
                sns.heatmap(mat, annot=True, fmt="d", cmap="YlGn",
                            xticklabels=all_em,
                            yticklabels=[h.split("-")[0].strip()[:20] for h in hod_ns],
                            linewidths=0.5, ax=ax, cbar=False)
                ax.set_title("(e) Emerging Hot Topics by HOD\n(1 = signal present in recent papers)",
                             fontsize=15, fontweight="bold")
                plt.xticks(rotation=40, ha="right", fontsize=11)
                plt.yticks(fontsize=11)
                plt.tight_layout(); st.pyplot(fig)

        # ── (f) Overlaps ───────────────────────────────────────────────────
        with q_tabs[5]:
            st.markdown('<div class="sec-header sh-purple">(f) Cross-HOD Research Overlaps</div>', unsafe_allow_html=True)
            hod_names = list(hod_keyword_sets.keys())
            n         = len(hod_names)
            overlap_records = []
            for i in range(n):
                for j in range(i+1, n):
                    common = sorted(hod_keyword_sets[hod_names[i]] & hod_keyword_sets[hod_names[j]])
                    if common:
                        overlap_records.append((hod_names[i], hod_names[j], common))

            if n >= 2:
                # Heatmap
                mat = np.zeros((n,n))
                for i in range(n):
                    for j in range(n):
                        if i==j: mat[i,j]=len(hod_keyword_sets[hod_names[i]])
                        else:     mat[i,j]=len(hod_keyword_sets[hod_names[i]] & hod_keyword_sets[hod_names[j]])
                sn = [h.split("-")[0].strip()[:16] for h in hod_names]
                fig, ax = plt.subplots(figsize=(max(7,n*1.8), max(6,n*1.5)))
                diag = np.eye(n,dtype=bool)
                sns.heatmap(mat, annot=True, fmt=".0f", cmap="Blues",
                            xticklabels=sn, yticklabels=sn,
                            linewidths=0.6, ax=ax, mask=diag, cbar_kws={"shrink":0.7})
                sns.heatmap(mat, annot=True, fmt=".0f", cmap="Greens",
                            xticklabels=sn, yticklabels=sn,
                            linewidths=0.6, ax=ax, mask=~diag, cbar=False)
                ax.set_title("(f) Cross-HOD Keyword Overlap\n(green=own; blue=shared)",
                             fontsize=15, fontweight="bold")
                plt.setp(ax.get_xticklabels(), rotation=30, ha="right", fontsize=10)
                plt.setp(ax.get_yticklabels(), fontsize=10)
                plt.tight_layout(); st.pyplot(fig)

                # Bubble matrix
                st.markdown('<div class="sec-header sh-purple">Collaboration Bubble Matrix</div>', unsafe_allow_html=True)
                lkp = {}
                for h1,h2,c in overlap_records:
                    lkp[(h1,h2)] = len(c); lkp[(h2,h1)] = len(c)

                mat2 = np.zeros((n,n))
                for i in range(n):
                    for j in range(n):
                        if i!=j: mat2[i,j] = lkp.get((hod_names[i],hod_names[j]),0)

                fig, ax = plt.subplots(figsize=(13,12))
                ax.set_facecolor("#F8F9FA")
                for k in range(n):
                    ax.axhline(k, color="#E0E0E0", lw=0.8, zorder=0)
                    ax.axvline(k, color="#E0E0E0", lw=0.8, zorder=0)

                def bcolor(v):
                    return "#1F3864" if v>=8 else "#2E75B6" if v>=5 else "#A8C4E0"

                mx = mat2.max() if mat2.max()>0 else 1
                for i in range(n):
                    for j in range(n):
                        v = mat2[i,j]
                        if v==0: continue
                        r = (v/mx)*0.42
                        ax.add_patch(plt.Circle((j,i),r,color=bcolor(v),zorder=3,lw=0))
                        ax.text(j,i,str(int(v)),ha="center",va="center",
                                fontsize=12,fontweight="bold",
                                color="white" if v>=5 else "#1F3864",zorder=4)
                for k in range(n):
                    ax.add_patch(plt.Rectangle((k-.5,k-.5),1,1,color="#F0F0F0",zorder=1))
                    ax.text(k,k,"—",ha="center",va="center",fontsize=12,color="#BBBBBB",zorder=2)

                ax.set_xlim(-.5,n-.5); ax.set_ylim(-.5,n-.5)
                ax.set_xticks(range(n)); ax.set_yticks(range(n))
                ax.set_xticklabels(sn, rotation=35, ha="right", fontsize=12)
                ax.set_yticklabels(sn, fontsize=12)
                ax.tick_params(length=0)
                ax.set_title("(f) Collaboration Potential — Bubble Matrix\nBubble size = shared keywords",
                             fontsize=16, fontweight="bold", pad=18)
                lp = [mpatches.Patch(color="#1F3864",label="High (≥8)"),
                      mpatches.Patch(color="#2E75B6",label="Medium (5–7)"),
                      mpatches.Patch(color="#A8C4E0",label="Low (1–4)")]
                ax.legend(handles=lp, fontsize=12, loc="lower right")
                ax.spines[["top","right","bottom","left"]].set_visible(False)
                plt.tight_layout(); st.pyplot(fig)

                # Table
                if overlap_records:
                    st.markdown('<div class="sec-header sh-purple">Top Collaboration Pairs</div>', unsafe_allow_html=True)
                    tbl = sorted(overlap_records, key=lambda x: len(x[2]), reverse=True)
                    df_tbl = pd.DataFrame([{
                        "HOD A": r[0].split("-")[0].strip()[:28],
                        "HOD B": r[1].split("-")[0].strip()[:28],
                        "Shared Keywords": len(r[2]),
                        "Top Keywords": ", ".join(r[2][:6]),
                        "Potential": "High" if len(r[2])>=8 else "Medium" if len(r[2])>=5 else "Emerging"
                    } for r in tbl])
                    st.dataframe(df_tbl, use_container_width=True, hide_index=True)


# ─────────────────────────────────────────────────────────────────────────────
# TAB 5 — EXPORT
# ─────────────────────────────────────────────────────────────────────────────
with tabs[5]:
    st.markdown('<div class="sec-header sh-green">💾 Export Results</div>', unsafe_allow_html=True)

    if not st.session_state.results:
        st.warning("Run NLP pipeline first.")
    else:
        preprocess = st.session_state.get("preprocess")
        if not preprocess:
            (STOP_WORDS, lemmatizer, word_tokenize, *_) = load_nlp_libs()
            preprocess = get_preprocess_fn(STOP_WORDS, lemmatizer, word_tokenize)
            st.session_state["preprocess"] = preprocess

        val_log = st.session_state.validation_log or {}

        if st.button("📊 Build & Download Excel Workbook", type="primary", use_container_width=True):
            with st.spinner("Building Excel workbook…"):
                xlsx_bytes = build_excel(st.session_state.results, val_log, preprocess)
            st.download_button(
                label="⬇️ Download UGBS_NLP_Analysis.xlsx",
                data=xlsx_bytes,
                file_name="UGBS_NLP_Analysis.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        st.markdown('<div class="sec-header sh-green">📄 Workbook Contents</div>', unsafe_allow_html=True)
        st.markdown("""
| Sheet | Contents |
|---|---|
| **Paper Summaries** | One row per paper — abstract · extractive summary · LDA keywords · BERTopic keywords · word count |
| **Accuracy Validation** | Keyword overlap % · ROUGE-1 F1 · hallucination flag · issues — for both models |
| **T5 Summaries** | All BART-generated abstractive summaries |
        """)


# ══════════════════════════════════════════════════════════════════════════════
#  FOOTER
# ══════════════════════════════════════════════════════════════════════════════
st.markdown("""
<div class="footer">
    🎓 UGBS Research NLP Analyser &nbsp;·&nbsp;
    Created by <strong>Jonathan Elikem</strong> &nbsp;·&nbsp;
    University of Ghana Business School &nbsp;·&nbsp; 2026
</div>""", unsafe_allow_html=True)
